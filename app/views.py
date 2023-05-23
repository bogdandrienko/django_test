from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
import json
import openpyxl


def ret_html(request):
    return render(request, "index.html")


def ret_str(request):
    return HttpResponse("1111111")


def ret_txt_file(request):
    with open("requirements.txt") as file:
        data = file.readlines()
    return HttpResponse(data)


def ret_json(request):
    data = [x for x in range(1, 100)]
    json_d = {"name": "Ali", "data": data}
    return JsonResponse(data=json_d, safe=False)


def ret_json_file(request):
    with open("data.json") as file:
        json_d = json.load(file)
    return JsonResponse(data=json_d, safe=False)


def ret_excel_file(request):
    workbook = openpyxl.load_workbook("data.xlsx")
    worksheet = workbook.active

    data = []
    for i in range(1, worksheet.max_row + 1):
        for j in range(1, worksheet.max_column + 1):
            data.append(worksheet.cell(i, j).value)
    return JsonResponse(data={"data": data}, safe=False)
